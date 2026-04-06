#!/usr/bin/env python3
"""
Alstom NDT — Skaner Prób v3.0
Hosting: Render.com
"""

from flask import Flask, request, jsonify, send_file, send_from_directory
from io import BytesIO
from datetime import date
import os
import json
import urllib.request
import urllib.error
import openpyxl
import smtplib
import base64
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

app = Flask(__name__, static_folder=".", static_url_path="")

@app.errorhandler(Exception)
def handle_exception(e):
    return jsonify({"error": str(e)}), 500

@app.errorhandler(404)
def handle_404(e):
    return send_from_directory(".", "index.html")

API_KEY      = os.environ.get("ANTHROPIC_API_KEY", "")
GMAIL_USER   = os.environ.get("GMAIL_USER", "")       # np. olgierd.moryc@gmail.com
GMAIL_PASS   = os.environ.get("GMAIL_APP_PASSWORD", "") # App Password z Google
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))

# Wczytaj szablony przy starcie
TEMPLATES = {}
for t in ["MT", "UT"]:
    path = os.path.join(BASE_DIR, f"{t}__wzor.xlsx")
    if os.path.exists(path):
        with open(path, "rb") as f:
            TEMPLATES[t] = f.read()
        print(f"✓ Wczytano {t}__wzor.xlsx")
    else:
        print(f"⚠ Brak {t}__wzor.xlsx!")


@app.route("/")
def index():
    return send_from_directory(".", "index.html")


@app.route("/api/scan", methods=["POST"])
def scan():
    body = request.get_json()
    image_b64 = body.get("image_base64", "")

    if not API_KEY:
        return jsonify({"error": "Brak klucza ANTHROPIC_API_KEY na serwerze"}), 400

    prompt = """Analizujesz formularz badań nieniszczących spawów firmy Alstom.

Odczytaj z nagłówka formularza:
1. SPAWACZ - tylko samo nazwisko (lub imię i nazwisko) bez tytułów i cyfr
2. PROJEKT - TYLKO sam kod/numer projektu z pierwszej ramki pola PROJEKT (np. "ALP", "ALP2") — ignoruj wszystko co jest napisane obok lub po prawej stronie (np. "PODŁUŻNICA DOLNA" to nie jest część projektu)

Następnie dla każdego wiersza tabeli z numerem próby (np. X01, X05, X13, O3, X31 itp.):
- Kolumna MT: wpis (MT+, MtΦ, symbol z plusem/kółkiem) I komórka NIE zacieniona → dodaj do mt_proby
- Kolumna UT: wpis (UT+, utΦ, symbol z plusem/kółkiem) I komórka NIE zacieniona → dodaj do ut_proby
- Komórki ZACIENIONE/SZARE = pomijamy

Zwróć WYŁĄCZNIE JSON, zero tekstu przed ani po:
{"spawacz":"...","projekt":"...","mt_proby":["X01"],"ut_proby":["X53"]}"""

    payload = json.dumps({
        "model": "claude-sonnet-4-6",
        "max_tokens": 1000,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": image_b64}},
                {"type": "text", "text": prompt}
            ]
        }]
    }).encode("utf-8")

    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=payload,
        headers={
            "Content-Type": "application/json",
            "x-api-key": API_KEY,
            "anthropic-version": "2023-06-01"
        }
    )

    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read())
        text = data["content"][0]["text"].strip().replace("```json", "").replace("```", "").strip()
        result = json.loads(text)
        return jsonify({"ok": True, "result": result})
    except urllib.error.HTTPError as e:
        try:
            err = json.loads(e.read()).get("error", {}).get("message", str(e))
        except:
            err = str(e)
        return jsonify({"error": err}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/export", methods=["POST"])
def export():
    body = request.get_json()
    t       = body.get("type", "MT").upper()
    proby   = body.get("proby", [])
    spawacz = body.get("spawacz", "")
    projekt = body.get("projekt", "")

    if t not in TEMPLATES:
        return jsonify({"error": f"Brak szablonu {t}__wzor.xlsx"}), 400

    wb = openpyxl.load_workbook(BytesIO(TEMPLATES[t]))
    ws = wb["DANE"]

    # Wyczyść tylko komórki danych (A1:B20), NIE dotykaj formuł w innych arkuszach
    for r in range(1, 21):
        ws.cell(row=r, column=1).value = None
        ws.cell(row=r, column=2).value = None

    # A1 = data (str1!I12)
    ws["A1"] = date.today()
    ws["A1"].number_format = "DD.MM.YYYY"

    # A2 = pierwsza próba / opis (str1!S17 — "PRÓBKA SPAWALNICZA")
    ws["A2"] = proby[0] if proby else ""

    # A3 = spawacz (str1!S21)
    ws["A3"] = spawacz

    # A4 = projekt (opcjonalnie, arkusze nie odwołują się, ale zachowujemy)
    ws["A4"] = projekt

    # B1..B20 = lista prób po jednej (str2!B29..B48)
    for i, p in enumerate(proby):
        if i >= 20:
            break
        ws.cell(row=i + 1, column=2).value = p

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Nazwa pliku: TYP_próba1_próba2_..._spawacz.xlsx
    spawacz_safe = spawacz.replace(" ", "_").upper()
    proby_str = "_".join(proby)
    fname = f"{t}_{proby_str}_{spawacz_safe}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname
    )


def build_xlsx(t, proby, spawacz, projekt):
    """Generuje plik XLSX w pamięci, zwraca (BytesIO, filename)."""
    wb = openpyxl.load_workbook(BytesIO(TEMPLATES[t]))
    ws = wb["DANE"]
    for r in range(1, 21):
        ws.cell(row=r, column=1).value = None
        ws.cell(row=r, column=2).value = None
    ws["A1"] = date.today()
    ws["A1"].number_format = "DD.MM.YYYY"
    ws["A2"] = proby[0] if proby else ""
    ws["A3"] = spawacz
    ws["A4"] = projekt
    for i, p in enumerate(proby):
        if i >= 20: break
        ws.cell(row=i + 1, column=2).value = p
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    spawacz_safe = (spawacz or "").replace(" ", "_").upper()
    proby_str = "_".join(proby)
    fname = f"{t}_{proby_str}_{spawacz_safe}.xlsx"
    return buf, fname


@app.route("/api/email", methods=["POST"])
def send_email():
    if not GMAIL_USER or not GMAIL_PASS:
        return jsonify({"error": "Brak konfiguracji Gmail na serwerze (GMAIL_USER / GMAIL_APP_PASSWORD)"}), 400

    body = request.get_json()
    to_addr = body.get("to", "").strip()
    mt_proby = body.get("mt_proby", [])
    ut_proby = body.get("ut_proby", [])
    spawacz  = body.get("spawacz", "")
    projekt  = body.get("projekt", "")

    if not to_addr:
        return jsonify({"error": "Brak adresu odbiorcy"}), 400

    # Treść maila HTML
    today_str = date.today().strftime("%d.%m.%Y")
    mt_list = "".join(f"<li>{p}</li>" for p in mt_proby) or "<li>brak</li>"
    ut_list = "".join(f"<li>{p}</li>" for p in ut_proby) or "<li>brak</li>"
    total = len(mt_proby) + len(ut_proby)

    html_body = f"""
<html><body style="font-family:Arial,sans-serif;color:#222;max-width:600px">
<div style="background:#0d0f14;padding:16px 24px;border-radius:8px 8px 0 0">
  <span style="font-family:monospace;color:#00c8ff;font-size:18px;font-weight:bold">NDT Skaner — Alstom</span>
</div>
<div style="border:1px solid #ddd;border-top:none;padding:24px;border-radius:0 0 8px 8px">
  <p>Dzień dobry,</p>
  <p>W załączeniu przesyłam wyniki badań nieniszczących z dnia <strong>{today_str}</strong>.</p>

  <table style="border-collapse:collapse;width:100%;margin:16px 0">
    <tr><td style="padding:8px;background:#f5f5f5;font-weight:bold;width:140px">Spawacz</td>
        <td style="padding:8px;border:1px solid #ddd">{spawacz}</td></tr>
    <tr><td style="padding:8px;background:#f5f5f5;font-weight:bold">Projekt</td>
        <td style="padding:8px;border:1px solid #ddd">{projekt}</td></tr>
    <tr><td style="padding:8px;background:#f5f5f5;font-weight:bold">Łącznie prób</td>
        <td style="padding:8px;border:1px solid #ddd">{total}</td></tr>
  </table>

  <table style="border-collapse:collapse;width:100%;margin:16px 0">
    <tr>
      <td style="width:50%;vertical-align:top;padding-right:12px">
        <div style="background:#fff3ef;border:1px solid #ff6b35;border-radius:6px;padding:12px">
          <strong style="color:#ff6b35">🔥 Próby MT ({len(mt_proby)})</strong>
          <ul style="margin:8px 0 0 0;padding-left:20px">{mt_list}</ul>
        </div>
      </td>
      <td style="width:50%;vertical-align:top;padding-left:12px">
        <div style="background:#edfaff;border:1px solid #00c8ff;border-radius:6px;padding:12px">
          <strong style="color:#0088aa">🔊 Próby UT ({len(ut_proby)})</strong>
          <ul style="margin:8px 0 0 0;padding-left:20px">{ut_list}</ul>
        </div>
      </td>
    </tr>
  </table>

  <p style="color:#666;font-size:13px">Pliki Excel z protokołami w załączniku.<br>
  Wygenerowano automatycznie przez system NDT Skaner.</p>
  <p>Pozdrawiam,<br><strong>{spawacz}</strong></p>
</div>
</body></html>
"""

    # Buduj wiadomość
    msg = MIMEMultipart("mixed")
    msg["From"]    = GMAIL_USER
    msg["To"]      = to_addr
    msg["Subject"] = f"NDT Wyniki — {projekt} — {spawacz} — {today_str}"

    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(html_body, "html", "utf-8"))
    msg.attach(alt)

    # Załącz pliki XLSX
    attached = []
    for t, proby in [("MT", mt_proby), ("UT", ut_proby)]:
        if not proby or t not in TEMPLATES:
            continue
        buf, fname = build_xlsx(t, proby, spawacz, projekt)
        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(buf.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", "attachment", filename=fname)
        msg.attach(part)
        attached.append(fname)

    if not attached:
        return jsonify({"error": "Brak prób do załączenia"}), 400

    # Wyślij przez Gmail SMTP
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(GMAIL_USER, GMAIL_PASS)
            smtp.sendmail(GMAIL_USER, to_addr, msg.as_bytes())
        return jsonify({"ok": True, "sent_to": to_addr, "files": attached})
    except smtplib.SMTPAuthenticationError:
        return jsonify({"error": "Błąd logowania Gmail — sprawdź GMAIL_APP_PASSWORD"}), 401
    except Exception as e:
        return jsonify({"error": f"Błąd wysyłki: {str(e)}"}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 3000))
    app.run(host="0.0.0.0", port=port)
